import requests
import json
import openpyxl
from datetime import datetime

FILE_NAME = 'Лизингополучатели.xlsx'
book = openpyxl.load_workbook(FILE_NAME)
sheet = book['Leasing']


def get_inns_from_excel():
    rows = iter(sheet['B'])
    next(rows)

    inns = []
    for row in rows:
        if row.value is not None:
            inns.append(row.value)
        else:
            break
    return inns


def get_guid_by_inn(inn: str):
    headers = {
        'X-KL-Ajax-Request': 'Ajax_Request',
        'Pragma': 'no-cache',
        'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
        'DNT': '1',
        'sec-ch-ua-mobile': '?0',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Cache-Control': 'no-cache',
        'Referer': f'https://fedresurs.ru/search/entity?code={inn}',
        'sec-ch-ua-platform': '"Windows"',
    }

    params = {
        'limit': '15',
        'offset': '0',
        'code': inn,
        'isActive': 'true',
    }

    response = requests.get('https://fedresurs.ru/backend/companies', params=params, headers=headers)
    data_json = json.loads(response.text)
    # print(data_json)
    guid = data_json['pageData'][0]['guid']
    # print(guid)
    return guid


def show_results():
    end_time = datetime.now()
    exec_time = end_time - start_time
    print("========================= FINISH =========================\n")
    print("Total time: ", exec_time)
    print("Orders checked: ", len(get_inns_from_excel()))


def get_orders_by_inn(guid):
    headers = {
        'X-KL-Ajax-Request': 'Ajax_Request',
        'Pragma': 'no-cache',
        'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
        'DNT': '1',
        'sec-ch-ua-mobile': '?0',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Cache-Control': 'no-cache',
        'Referer': f'https://fedresurs.ru/company/{guid}',
        'sec-ch-ua-platform': '"Windows"',
    }

    params = {
        'limit': '50',
        'offset': '0',
        'searchCompanyEfrsb': 'true',
        'searchAmReport': 'true',
        'searchFirmBankruptMessage': 'true',
        'searchFirmBankruptMessageWithoutLegalCase': 'false',
        'searchSfactsMessage': 'true',
        'searchSroAmMessage': 'true',
        'searchTradeOrgMessage': 'true',
    }

    response = requests.get(
        f'https://fedresurs.ru/backend/companies/{guid}/publications',
        params=params,
        headers=headers,
    )

    data_json = json.loads(response.text)
    return data_json


def get_order_data(order_link):
    """
    returns json
    """
    headers = {
        'X-KL-Ajax-Request': 'Ajax_Request',
        'Pragma': 'no-cache',
        'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
        'DNT': '1',
        'sec-ch-ua-mobile': '?0',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Cache-Control': 'no-cache',
        'Referer': f'https://fedresurs.ru/sfactmessage/{order_link}',
        'sec-ch-ua-platform': '"Windows"',
    }

    order_response = requests.get(f'https://fedresurs.ru/backend/sfactmessages/{order_link}', headers=headers)
    order_json = json.loads(order_response.text)
    return order_json


def insert(row, col, data):
    sheet.cell(row=row, column=col).value = data


def get_data(row, col):
    """
    :param row:
    :param col:
    :return: value from the cell
    """
    return sheet.cell(row=row, column=col).value


def insert_order_data(row, order_data, is_first, orders_amount):
    """
    get order data and insert to the cell
    """

    if is_first:
        company_name = get_data(row, 1)
        inn_number = get_data(row, 2)
    else:
        company_name = get_data(row-1, 1)
        inn_number = get_data(row-1, 2)

    order_type = order_data['typeName']

    insert(row, 1, company_name)
    insert(row, 2, inn_number)
    insert(row, 9, orders_amount)
    insert(row, 10, order_type)  # order name

    try:
        lessor_company = order_data['content']['lessorsCompanies'][0]['fullName']
    except KeyError:
        lessor_company = ""

    if 'startDate' in order_data['content']:
        try:
            obj = order_data['content']['subjects'][0]['description']
        except KeyError:
            obj = ""

        publisher = order_data['publisher']['name']

        insert(row, 5, obj)
        insert(row, 6, lessor_company)
        insert(row, 8, publisher)

        start_date = order_data['content']['startDate'].split('T')[0]
        end_date = order_data['content']['endDate'].split('T')[0]

        insert(row, 3, start_date)
        insert(row, 4, end_date)
    if 'stopDate' in order_data['content']:
        publisher = order_data['publisher']['name']
        stop_date = order_data['content']['stopDate'].split('T')[0]

        insert(row, 6, lessor_company)
        insert(row, 8, publisher)
        insert(row, 7, stop_date)


def save_and_close(book):
    book.save(FILE_NAME)
    book.close()


def fill_notations(current_row, orders_amount):
    notations = []
    for row in range(current_row - orders_amount, current_row):
        notation = ''
        for col in range(3, 10):
            if get_data(row, col) is not None:
                notation += str(get_data(row, col)) + ' / '
        notations.append(notation)

    index = 0
    for row in range(current_row - orders_amount, current_row):
        for col in range(11, 11 + orders_amount):
            insert(row, col, notations[index])
            index += 1
        index = 0


def main():

    current_row = 2

    for inn in get_inns_from_excel():

        is_first_order = True

        print("Current INN: ", inn)
        guid = get_guid_by_inn(inn)
        orders = get_orders_by_inn(guid)

        orders_amount = len(orders['pageData'])
        print('Orders amount: ', orders_amount)

        orders_links = []
        for order in orders['pageData']:
            print('Order: ', order['title'])
            orders_links.append(order['guid'])

        if orders_amount > 0:
            for order_link in orders_links:
                order_data = get_order_data(order_link)
                if is_first_order:
                    insert_order_data(current_row, order_data, is_first_order, orders_amount)
                    is_first_order = False
                else:
                    sheet.insert_rows(current_row)
                    insert_order_data(current_row, order_data, is_first_order, orders_amount)
                current_row += 1

            fill_notations(current_row, orders_amount)
        else:
            current_row += 1

    print("=============================================================")

    save_and_close(book)


if __name__ == "__main__":
    start_time = datetime.now()
    main()
    show_results()
